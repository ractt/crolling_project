import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import requests
from bs4 import BeautifulSoup
from collections import Counter
import re
import pyperclip
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl


# 엑셀 파일로 저장하는 함수
def save_excel_file(filename, data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # 데이터를 엑셀 파일에 쓰기
    for row, result in enumerate(data, start=1):
        product, count = result.split()
        sheet.cell(row=row, column=1).value = product
        sheet.cell(row=row, column=2).value = int(count)
    
    # 파일 저장
    wb.save(filename)

# 네이버 쇼핑 검색 함수
def search_naver(keyword, browser):
    browser.get(f'https://search.shopping.naver.com/search/all?query={keyword}')
    time.sleep(2)
    before_h_naver = browser.execute_script("return window.scrollY")
    result_naver_raw = []

    while True:
        browser.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)
        time.sleep(1)
        after_h_naver = browser.execute_script("return window.scrollY")

        if after_h_naver == before_h_naver:
            break

        before_h_naver = after_h_naver

    content_naver = browser.find_elements(By.CLASS_NAME, 'product_title__Mmw2K')
    for element_naver in content_naver:
        result_naver_raw.append(element_naver.text)

    return result_naver_raw

# 11번가 검색 함수
def search_11st(keyword, browser):
    browser.get(f'https://search.11st.co.kr/pc/total-search?kwd={keyword}&tabId=TOTAL_SEARCH')
    time.sleep(2)
    result_11st_raw = []

    commonPrd_11st = browser.find_element(By.XPATH, "//*[@id='section_commonPrd']/div[2]/ul")
    content_11st = commonPrd_11st.find_elements(By.CLASS_NAME, "c-card-item__name")
    
    for element_11st_raw in content_11st:
        element_11st_text = element_11st_raw.text
        if "상품명" in element_11st_text:
            element_11st_text_non_productName = element_11st_text.split("상품명")[1].strip() 
            result_11st_raw.append(element_11st_text_non_productName)  # 상품명을 리스트에 추가

    return result_11st_raw 

# 지마켓 검색 함수
def search_gmarket(keyword):
    url_gmarket = f"https://browse.gmarket.co.kr/search?keyword={keyword}"
    response_gmarket = requests.get(url_gmarket)
    soup_gmarket = BeautifulSoup(response_gmarket.text, 'html.parser')
    items_gmarket_text_selected = soup_gmarket.select(".text__item")
    results_gmarket_raw = [item_gmarket_element.text.strip() for item_gmarket_element in items_gmarket_text_selected]
    return results_gmarket_raw

# 네이버 검색 결과를 엑셀 파일로 저장하는 함수
def save_naver_to_excel():
    global naver_results
    save_excel_file("네이버_검색결과.xlsx", naver_results)

# 11번가 검색 결과를 엑셀 파일로 저장하는 함수
def save_11st_to_excel():
    global eleven_street_results
    save_excel_file("11번가_검색결과.xlsx", eleven_street_results)

# 지마켓 검색 결과를 엑셀 파일로 저장하는 함수
def save_gmarket_to_excel():
    global gmarket_results
    save_excel_file("지마켓_검색결과.xlsx", gmarket_results)

# 검색 함수
def search(event=None):
    keyword = entry.get()
    chrome_options = Options()
    chrome_options.add_argument("--incognito")
    browser = webdriver.Chrome(options=chrome_options)

    try:
        # 포셀 사이트 로그인 및 검색
        browser.get('http://forsell.co.kr/login/')
        time.sleep(2)
        id_input_forsell = browser.find_element(By.NAME, 'username')
        pw_input_forsell = browser.find_element(By.NAME, 'password')
        id_input_forsell.send_keys('aaa10130')
        pw_input_forsell.send_keys('tjgktjd123')
        login_button_forsell = browser.find_element(By.CSS_SELECTOR, 'button[type="submit"]')
        login_button_forsell.click()
        time.sleep(5)
        SearchTab_forsell = browser.find_element(By.LINK_TEXT, '연관키워드 검색')
        SearchTab_forsell.click()
        time.sleep(1)

        try:
            # 광고 닫기
            action_offad_forsell = ActionChains(browser)
            action_offad_forsell.move_by_offset(100, 100).click().perform()
        except NoSuchElementException:
            print("닫기 없음________________________________________")
            pass

        time.sleep(4)
        keyword_input_forsell = browser.find_element(By.ID, 'searchKeyword')
        keyword_input_forsell.send_keys(keyword)
        search_button_forsell = browser.find_element(By.ID, 'searchBtn')
        search_button_forsell.click()
        time.sleep(7)
        coupang1_inforsell = browser.find_element(By.ID, 'coupangReKeyword')
        coupang2_inforsell = browser.find_element(By.ID, 'coupangAutoKeyword')
        coupang1_inforsell.click()
        coupang2_inforsell.click()
        time.sleep(2)
        copy_forsell_keyword = browser.find_element(By.ID, 'copyKeyword')
        copy_forsell_keyword.click()
        clipboard_content_forsell = pyperclip.paste()
        result_forsell = clipboard_content_forsell.split(',')

        # 네이버 쇼핑 검색
        result_naver_raw = search_naver(keyword, browser)
        
        # 11번가 검색
        result_11st_raw = search_11st(keyword, browser)

        # 지마켓 검색
        results_gmarket_raw = search_gmarket(keyword)

        # 검색 결과 저장
        global naver_results, eleven_street_results, gmarket_results
        naver_results = result_naver_raw
        eleven_street_results = result_11st_raw
        gmarket_results = results_gmarket_raw
        
        # 검색 결과 출력
        result_text_naver.delete(1.0, tk.END)
        result_text_11st.delete(1.0, tk.END)
        result_text_gmarket.delete(1.0, tk.END)
        
        for product_name, count in naver_results:
            product_name = product_name.strip("(),'") 
            result_text_naver.insert(tk.END, f"{product_name} {count}\n")
        for product_name, count in eleven_street_results:
            product_name = product_name.strip("(),'") 
            result_text_11st.insert(tk.END, f"{product_name} {count}\n")
        for product_name, count in gmarket_results:
            product_name = product_name.strip("(),'") 
            result_text_gmarket.insert(tk.END, f"{product_name} {count}\n")

    finally:
        browser.quit()


# 클립보드에 텍스트 복사하는 함수
def copy_to_clipboard(text):
    pyperclip.copy(text)

# 엑셀 버튼을 눌렀을 때 엑셀 파일로 저장하는 함수
def save_to_excel():
    global naver_results, eleven_street_results, gmarket_results
    all_results = naver_results + eleven_street_results + gmarket_results
    save_excel_file("검색결과.xlsx", all_results)

# Tkinter 윈도우 생성
root = tk.Tk()
root.title("검색 앱")
root.geometry("2000x800")

# 라벨과 엔트리 위젯 생성
label = ttk.Label(root, text="검색어를 입력하세요:")
label.grid(row=0, column=0, padx=5, pady=5)

entry = ttk.Entry(root, width=30)
entry.grid(row=0, column=1, padx=5, pady=5)

# 검색 버튼 생성
search_button = ttk.Button(root, text="검색", command=search)
search_button.grid(row=0, column=2, padx=5, pady=5)

# 네이버 검색 결과를 표시할 텍스트 박스 생성
result_text_naver = tk.Text(root, height=20, width=40)
result_text_naver.grid(row=1, column=0, padx=5, pady=5)

# 11번가 검색 결과를 표시할 텍스트 박스 생성
result_text_11st = tk.Text(root, height=20, width=40)
result_text_11st.grid(row=1, column=1, padx=5, pady=5)

# 지마켓 검색 결과를 표시할 텍스트 박스 생성
result_text_gmarket = tk.Text(root, height=20, width=40)
result_text_gmarket.grid(row=1, column=2, padx=5, pady=5)

# 포셀 검색 결과를 표시할 텍스트 박스 생성
result_text_forsell = tk.Text(root, height=20, width=40)
result_text_forsell.grid(row=1, column=3, padx=5, pady=5)

# 텍스트의 글자 크기 설정
result_text_naver.config(font=("Courier", 15))
result_text_11st.config(font=("Courier", 15))
result_text_gmarket.config(font=("Courier", 15))
result_text_forsell.config(font=("Courier", 15))

# 네이버 검색 결과 복사 버튼
copy_naver_button = ttk.Button(root, text="네이버 결과 복사", command=lambda: copy_to_clipboard(result_text_naver.get(1.0, tk.END)))
copy_naver_button.grid(row=2, column=0, padx=5, pady=5)
copy_11st_button = ttk.Button(root, text="11번가 결과 복사", command=lambda: copy_to_clipboard(result_text_11st.get(1.0, tk.END)))
copy_11st_button.grid(row=2, column=1, padx=5, pady=5)
copy_gmarket_button = ttk.Button(root, text="지마켓 결과 복사", command=lambda: copy_to_clipboard(result_text_gmarket.get(1.0, tk.END)))
copy_gmarket_button.grid(row=2, column=2, padx=5, pady=5)
copy_forsell_button = ttk.Button(root, text="포셀 결과 복사", command=lambda: copy_to_clipboard(result_text_forsell.get(1.0, tk.END)))
copy_forsell_button.grid(row=2, column=3, padx=5, pady=5)

#엑셀 버튼
# 각 사이트별 엑셀 버튼 생성
excel_button_naver = ttk.Button(root, text="네이버 검색 결과 엑셀로 저장", command=save_naver_to_excel)
excel_button_naver.grid(row=2, column=0, padx=5, pady=5)
excel_button_11st = ttk.Button(root, text="11번가 검색 결과 엑셀로 저장", command=save_11st_to_excel)
excel_button_11st.grid(row=2, column=1, padx=5, pady=5)
excel_button_gmarket = ttk.Button(root, text="지마켓 검색 결과 엑셀로 저장", command=save_gmarket_to_excel)
excel_button_gmarket.grid(row=2, column=2, padx=5, pady=5)
forsell_button = ttk.Button(root, text="포셀 엑셀", command=lambda: None)
forsell_button.grid(row=3, column=3, padx=5, pady=5)


# 엔터 키를 누르면 검색 버튼 클릭
root.bind('<Return>', search)

# GUI 실행
root.mainloop()
