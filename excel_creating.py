import openpyxl

def save_excel_file(filename, data):
    # 엑셀 파일 생성
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # 데이터를 엑셀 파일에 쓰기
    for row, result in enumerate(data, start=1):
        product, count = result.split()
        sheet.cell(row=row, column=1).value = product
        sheet.cell(row=row, column=2).value = int(count)
    
    # 파일 저장
    wb.save(filename)

# 검색 결과 데이터
search_results = [
    "가위 26", "고기 12", "주방가위 12", "주방 11", "올스텐 10", "분리형 8", "식가위 5", "스텐 5", "다용도 5",
    "캠핑가위 4", "이유식 4", "업소용 3", "캠핑용 3", "벨가위 3", "일체형 3", "휴대용 3", "세라믹 3", "즈윌링 3",
    "잘드는 2", "식당 2", "국산 2", "스테인레스 2", "고기가위 2", "드라이 2", "집게 2", "세트 2", "트윈 2",
    "이유식가위 2", "무디타 2", "폼폼 2", "스텐가위 2", "마이셰프 2", "캠핑 2", "피스카스 2"
]

# 함수 호출하여 엑셀 파일 생성
save_excel_file("검색결과.xlsx", search_results)
