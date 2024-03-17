import openpyxl

# 엑셀 파일 생성 함수
def create_excel(search_results, filename):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # 데이터를 엑셀 파일에 쓰기
    for row, result in enumerate(search_results, start=1):
        product, count = result.split()
        sheet.cell(row=row, column=1).value = product
        sheet.cell(row=row, column=2).value = int(count)

    # 파일 저장
    wb.save(filename)

# 네이버 검색 결과를 엑셀 파일로 저장하는 함수
def save_results_to_excel(text):
    search_results = text.split('\n')
    create_excel(search_results, "네이버_검색결과.xlsx")

# Tkinter 버튼 클릭 시 호출될 함수
def save_results():
    text = result_naver.get(1.0, tk.END)
    save_results_to_excel(text)

# Tkinter 버튼 생성
naver_button = ttk.Button(root, text="네이버 엑셀", command=save_naver_results)
naver_button.grid(row=3, column=0, padx=5, pady=5)
